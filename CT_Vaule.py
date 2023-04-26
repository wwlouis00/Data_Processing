import openpyxl
import time
import datetime
from datetime import datetime, timedelta
import statistics
import numpy as np
from openpyxl import load_workbook
from heapq import nsmallest
from line_profiler import LineProfiler
 
#存檔為現在輸出的時間
now_output_time = "Work"+str(datetime.now().strftime('%Y-%m-%d %H-%M-%S'))+".csv"
# 讀取 Excel 檔案
wb = openpyxl.load_workbook('work.xlsx')
ws = wb.active

# 設置IFC值
IFC1 = 116
IFC2 = 130
IFC5 = 142
IFC6 = 148

# 設標頭欄位
ws.cell(row=1, column=19, value= "Normalize" )
ws.cell(row=1, column=20, value= "Time")
ws.cell(row=1, column=21, value= "Period(sec)")

#Set_time()設置Normalize 
def Set_time():
    #設置well欄位
    for time in range(2,79,1):
        ws.cell(row = time,column = 20).value = ws.cell(row=time,column=1).value

    #累積時間欄位
    orgin_time = str(ws.cell(row = 2,column=20).value)
    d1 = datetime.strptime(orgin_time,"%H:%M:%S")
    for time in range(2,79,1):
        a = str(ws.cell(row= time,column=20).value)
        accumulation = datetime.strptime(a,"%H:%M:%S")
        delta = accumulation - d1
        ws.cell(row =time,column=21).value = delta.seconds /60
# 
def well():
    for well_row in range(2,18,1):
        ws.cell(row=1,column=well_row+20 ,value = "well" + str(well_row-1))#設好"Moving"對應欄位
        ws.cell(row=1,column=well_row+40 ,value = "well" + str(well_row-1))#設好"Normalize"對應欄位
        #平均
        for Awell_row in range(2,79,1):
            gum = ws.cell(row=3,column=well_row).value
            for i in range(3,14,1):
                 gum = gum + ws.cell(row=i+1,column=well_row).value
            avg = float(gum/12)
            a = ws.cell(row = Awell_row,column = well_row).value
            threshold = float(a - avg)/IFC1
            ws.cell(row = Awell_row,column = well_row + 20).value = threshold  
        
        #算Ct_threshold
        CT_threshold = []    
        for mwell_row in range(3,15,1):
            well_gap = float(ws.cell(row = mwell_row,column= well_row +20).value)
            CT_threshold.append(well_gap)

        std = float(statistics.stdev(CT_threshold))
        aum = float(ws.cell(row=3,column= well_row + 20).value)

        for avg_row in range(3,14,1):
            aum = aum + ws.cell(row= avg_row+1,column= well_row +20).value
        aum_avg = aum / 10
        ws.cell(row = 2,column=well_row + 40).value = 10*std + aum_avg
        
        #進行切線計算
        Everywell_th = []
        for k in range(2,75,1): 
            a = float(ws.cell(row = k,column=well_row + 20).value)
            Everywell_th.append(a)
        b = float(ws.cell(row = 2,column=well_row + 40).value) # threshold
        c = min(Everywell_th, key=lambda x:(x<b)) #最接近threshold的數字
        well_threshold_row = Everywell_th.index(c)
        o = well_threshold_row
        th1 = float(ws.cell(row = o+2 ,column= well_row +20).value)
        time_1 = ws.cell(row = o+2,column= 21).value
        th2 = float(ws.cell(row = o+1,column= well_row +20).value)
        time_2 = ws.cell(row = o+1,column= 21).value
        high = th1- th2
        gap_time = float(time_1 - time_2) #三角形底 秒數
        high_2 = float(b - th2)
        high_gap = high / high_2
        gap_time_2 = gap_time / high_gap
        CT_final= gap_time_2 + time_2
        ws.cell(row = 10,column=well_row+40).value = CT_final

def main():
    Set_time()
    well()

main()
wb.save(now_output_time)
